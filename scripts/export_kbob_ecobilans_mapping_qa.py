#!/usr/bin/env python3
"""Export a local QA spreadsheet: abstract product names ↔ KBOB Baumaterialien row labels.

Reads product→KBOB mapping TTL files and resolves KBOB row names from a local JSON extract
(temp/kbob-ecobilans-construction-v8.02.json). Output stays under temp/ (gitignored);
KBOB indicator data and labels are not published from this repository.

Usage:
  python scripts/export_kbob_ecobilans_mapping_qa.py
  python scripts/export_kbob_ecobilans_mapping_qa.py --output temp/my-qa.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rdflib import Graph, URIRef
from rdflib.namespace import SKOS

REPO_ROOT = Path(__file__).resolve().parents[1]
CLASSIFICATION_DIR = REPO_ROOT / "classification"
CATALOG_PATH = CLASSIFICATION_DIR / "catalog.yaml"
DEFAULT_JSON = REPO_ROOT / "temp/kbob-ecobilans-construction-v8.02.json"
DEFAULT_OUTPUT = REPO_ROOT / "temp/kbob-ecobilans-mapping-qa.xlsx"

KBOB_MAPPING_SLUG_SUFFIX = "-to-kbob-ecobilans"
MATCH_PREDICATES = (SKOS.closeMatch, SKOS.relatedMatch)
KBOB_ID_PATTERN = re.compile(r"^\d{2}\.\d{3}$")

HEADERS = [
    "mapping",
    "product_notation",
    "product_name_en",
    "product_name_de",
    "relation",
    "kbob_id",
    "kbob_name_de",
    "kbob_name_fr",
    "kbob_ref_unit",
    "kbob_id_found",
]


def load_catalog() -> tuple[dict[str, str], dict[str, str]]:
    """Return vocab slug→path and mapping slug→path from catalog.yaml."""
    data = yaml.safe_load(CATALOG_PATH.read_text(encoding="utf-8"))
    vocab_paths = {entry["slug"]: entry["path"] for entry in data.get("vocabularies", [])}
    mapping_paths = {entry["slug"]: entry["path"] for entry in data.get("mappings", [])}
    return vocab_paths, mapping_paths


def vocab_slug_from_mapping_slug(mapping_slug: str) -> str:
    base = mapping_slug[: -len(KBOB_MAPPING_SLUG_SUFFIX)]
    if base.endswith("-products"):
        return f"{base[:-1]}"  # ...-products → ...-product
    return base


def kbob_mapping_pairs(
    mapping_paths: dict[str, str],
    vocab_paths: dict[str, str],
) -> list[tuple[str, str, str]]:
    """(mapping_slug, mapping_path, vocab_slug) for each product→KBOB bridge."""
    pairs: list[tuple[str, str, str]] = []
    for slug, path in sorted(mapping_paths.items()):
        if not slug.endswith(KBOB_MAPPING_SLUG_SUFFIX):
            continue
        if slug.startswith("kbob-"):
            continue
        vocab_slug = vocab_slug_from_mapping_slug(slug)
        if vocab_slug not in vocab_paths:
            print(f"Warning: no vocabulary slug {vocab_slug!r} for mapping {slug}", file=sys.stderr)
            continue
        pairs.append((slug, path, vocab_slug))
    return pairs


def local_name(term: URIRef) -> str:
    text = str(term)
    if "#" in text:
        return text.rsplit("#", 1)[-1]
    return text.rstrip("/").rsplit("/", 1)[-1]


def kbob_id_from_term(graph: Graph, term: URIRef) -> str:
    notation = graph.value(term, SKOS.notation)
    if notation is not None:
        return str(notation)
    local = local_name(term)
    if "_" in local:
        group, rest = local.split("_", 1)
        if group.isdigit() and rest.isdigit():
            return f"{group}.{rest}"
    return local.replace("_", ".")


def load_product_labels(vocab_path: Path) -> dict[str, dict[str, str]]:
    graph = Graph()
    graph.parse(vocab_path, format="turtle")
    labels: dict[str, dict[str, str]] = {}
    for concept in graph.subjects(SKOS.notation, None):
        notation = str(graph.value(concept, SKOS.notation))
        names: dict[str, str] = {}
        for _, _, label in graph.triples((concept, SKOS.prefLabel, None)):
            lang = label.language or "und"
            names[lang] = str(label)
        labels[notation] = names
    return labels


def load_kbob_rows(json_path: Path) -> dict[str, dict[str, str | None]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    rows = data["sheets"]["baumaterialien"]["rows"]
    by_id: dict[str, dict[str, str | None]] = {}
    for row in rows:
        if row.get("row_type") != "entry":
            continue
        kbob_id = row["id"]
        if not KBOB_ID_PATTERN.match(kbob_id):
            continue
        by_id[kbob_id] = {
            "label_de": row.get("label_de"),
            "label_fr": row.get("label_fr"),
            "ref_unit": row.get("ref_unit"),
        }
    return by_id


def parse_mapping_links(mapping_path: Path) -> list[tuple[str, str, str]]:
    """Return (product_notation, relation, kbob_id) for each close/related match."""
    graph = Graph()
    graph.parse(mapping_path, format="turtle")

    product_notation: dict[URIRef, str] = {}
    for concept in set(subj for subj, pred, _ in graph.triples((None, None, None)) if pred in MATCH_PREDICATES):
        notation = graph.value(concept, SKOS.notation)
        if notation is not None:
            product_notation[concept] = str(notation)
        else:
            product_notation[concept] = local_name(concept)

    links: list[tuple[str, str, str]] = []
    for subj, pred, obj in graph.triples((None, None, None)):
        if pred not in MATCH_PREDICATES or not isinstance(subj, URIRef) or not isinstance(obj, URIRef):
            continue
        relation = local_name(pred)
        source = product_notation.get(subj, local_name(subj))
        target = kbob_id_from_term(graph, obj)
        links.append((source, relation, target))

    links.sort(key=lambda item: (item[0], item[1], item[2]))
    return links


def build_rows(
    mapping_slug: str,
    mapping_rel_path: str,
    vocab_rel_path: str,
    kbob_by_id: dict[str, dict[str, str | None]],
) -> list[list[object]]:
    vocab_path = CLASSIFICATION_DIR / vocab_rel_path
    mapping_path = CLASSIFICATION_DIR / mapping_rel_path
    if not vocab_path.is_file():
        raise FileNotFoundError(f"Vocabulary not found: {vocab_path}")
    if not mapping_path.is_file():
        raise FileNotFoundError(f"Mapping not found: {mapping_path}")

    product_labels = load_product_labels(vocab_path)
    links = parse_mapping_links(mapping_path)
    rows: list[list[object]] = []

    for product_notation, relation, kbob_id in links:
        names = product_labels.get(product_notation, {})
        kbob = kbob_by_id.get(kbob_id, {})
        rows.append(
            [
                mapping_slug,
                product_notation,
                names.get("en", ""),
                names.get("de", ""),
                relation,
                kbob_id,
                kbob.get("label_de") or "",
                kbob.get("label_fr") or "",
                kbob.get("ref_unit") or "",
                kbob_id in kbob_by_id,
            ]
        )
    return rows


def autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            value = row[0].value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def write_workbook(rows: list[list[object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mapping_qa"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)

    missing = [row for row in rows if not row[9]]
    if missing:
        ws_missing = wb.create_sheet("missing_kbob_ids")
        ws_missing.append(HEADERS)
        for cell in ws_missing[1]:
            cell.font = Font(bold=True)
        for row in missing:
            ws_missing.append(row)
        autosize_columns(ws_missing)

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--json",
        type=Path,
        default=DEFAULT_JSON,
        help=f"Local KBOB JSON extract (default: {DEFAULT_JSON.relative_to(REPO_ROOT)})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output Excel path (default: {DEFAULT_OUTPUT.relative_to(REPO_ROOT)})",
    )
    args = parser.parse_args()

    if not args.json.is_file():
        print(
            f"KBOB JSON not found: {args.json}\n"
            "Run: python scripts/convert_kbob_ecobilans_xlsx_to_json.py",
            file=sys.stderr,
        )
        return 1

    vocab_paths, mapping_paths = load_catalog()
    kbob_by_id = load_kbob_rows(args.json)

    all_rows: list[list[object]] = []
    for mapping_slug, mapping_rel, vocab_slug in kbob_mapping_pairs(mapping_paths, vocab_paths):
        vocab_rel = vocab_paths[vocab_slug]
        all_rows.extend(build_rows(mapping_slug, mapping_rel, vocab_rel, kbob_by_id))

    all_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[4]), str(r[5])))
    write_workbook(all_rows, args.output)

    missing_count = sum(1 for row in all_rows if not row[9])
    print(f"Wrote {len(all_rows)} mapping links to {args.output}")
    if missing_count:
        print(f"Warning: {missing_count} KBOB id(s) not found in JSON", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
