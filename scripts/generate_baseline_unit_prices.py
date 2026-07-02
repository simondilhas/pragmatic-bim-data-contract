#!/usr/bin/env python3
"""Generate baseline unit price JSON with KBOB ecobilans carbon data from Excel and mapping TTL."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
from rdflib import Graph
from rdflib.namespace import SKOS

from baseline_prices_common import (
    ANCHOR_REGION,
    DISCLAIMER,
    LAYER_CATEGORY_BY_PREFIX,
    LAYER_NOTATION_SUFFIXES,
    LICENSE,
    NOTATION_PREFIX_CATEGORIES,
    SCHEME_IRI_PREFIXES,
    build_uncertainty,
    format_demolition,
    infer_product_category,
    normalize_unit,
    notation_to_iri,
    parse_float,
    resolve_provenance_status,
    resolve_uncertainty_config,
    scale_bounds,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = REPO_ROOT / "baseline-unit-prices/PragmaticBIM_Strict_SKOS_Unified_Prices-v12.xlsx"
OUT_PRODUCTS = REPO_ROOT / "baseline-unit-prices/baseline-products.json"
OUT_REGIONAL = REPO_ROOT / "baseline-unit-prices/regional-benchmarks.json"
OUT_REGIONAL_LABOR = REPO_ROOT / "baseline-unit-prices/regional-labor-benchmarks.json"
OUT_REGIONAL_MATERIAL = REPO_ROOT / "baseline-unit-prices/regional-material-benchmarks.json"
COST_ADJUSTMENTS_PATH = REPO_ROOT / "baseline-unit-prices/cost-adjustments.json"
LABOR_OVERRIDES_PATH = REPO_ROOT / "baseline-unit-prices/labor-overrides.json"
REGIONAL_MATERIAL_PATH = REPO_ROOT / "baseline-unit-prices/regional-material-benchmarks.json"
MAPPING_GLOB = "abstract-*-to-kbob-ecobilans.mapping.ttl"
LCA_FACTORS_PATH = REPO_ROOT / "classification/mapping/kbob-ecobilans-lca-factors.json"
LAYER_RECIPES_PATH = REPO_ROOT / "classification/mapping/kbob-ecobilans-layer-recipes.json"
UNIT_DEFAULTS_PATH = REPO_ROOT / "classification/mapping/kbob-ecobilans-price-unit-defaults.json"
KBOB_SKOS_PATH = REPO_ROOT / "classification/mapping/kbob-ecobilans-baumat.skos.ttl"

KBOB_SCHEME = "https://example.org/ch/kbob/ecobilans/baumat/2009-1-2022/"

NOTATION_ALIASES: dict[str, str] = {}


def kbob_local_to_id(local: str) -> str:
    return local.replace("_", ".")


def kbob_id_to_iri(kbob_id: str) -> str:
    return KBOB_SCHEME + kbob_id.replace(".", "_")


def load_kbob_labels() -> dict[str, str]:
    labels: dict[str, str] = {}
    graph = Graph()
    graph.parse(KBOB_SKOS_PATH, format="turtle")
    for concept in graph.subjects(SKOS.notation, None):
        notation = str(graph.value(concept, SKOS.notation))
        label = graph.value(concept, SKOS.prefLabel, None)
        if notation and label:
            labels[notation] = str(label)
    return labels


def load_mappings() -> dict[str, dict[str, list[str]]]:
    mappings: dict[str, dict[str, list[str]]] = {}
    mapping_dir = REPO_ROOT / "classification/mapping"
    for path in sorted(mapping_dir.glob(MAPPING_GLOB)):
        graph = Graph()
        graph.parse(path, format="turtle")
        for subject, predicate, obj in graph.triples((None, None, None)):
            pred_str = str(predicate)
            if not (pred_str.endswith("closeMatch") or pred_str.endswith("relatedMatch")):
                continue
            match_type = "closeMatch" if pred_str.endswith("closeMatch") else "relatedMatch"
            subject_str = str(subject)
            notation = subject_str.rsplit("/", 1)[-1]
            kbob_local = str(obj).rsplit("/", 1)[-1]
            if not re.match(r"^\d{2}_\d{3}$", kbob_local):
                continue
            kbob_id = kbob_local_to_id(kbob_local)
            entry = mappings.setdefault(notation, {"closeMatch": [], "relatedMatch": []})
            if kbob_id not in entry[match_type]:
                entry[match_type].append(kbob_id)
    return mappings


def load_excel_products() -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Baseline Reference Data"]
    rows = list(ws.iter_rows(values_only=True))
    products: list[dict[str, object]] = []
    for row in rows[1:]:
        notation = row[0]
        if not notation:
            continue
        products.append(
            {
                "notation": str(notation).strip(),
                "label_en": str(row[1] or "").strip(),
                "label_de": str(row[2] or "").strip(),
                "material_cost": parse_float(row[3]),
                "unit": normalize_unit(str(row[4] or "")),
                "labor_hours": parse_float(row[5]),
                "category": str(row[6] or "").strip(),
            }
        )
    return products


def load_excel_regions() -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Regional Labor Settings"]
    rows = list(ws.iter_rows(values_only=True))
    regions: list[dict[str, object]] = []
    for row in rows[1:]:
        code = row[0]
        if not code:
            continue
        regions.append(
            {
                "code": str(code).strip(),
                "label_en": str(row[1] or "").strip(),
                "hourly_rate": parse_float(row[2]),
                "currency": str(row[3] or "").strip(),
            }
        )
    return regions


def build_price_labor_demolition(
    product: dict[str, object],
    cost_adjustments: dict,
    labor_overrides: dict[str, dict],
) -> tuple[dict, dict, dict]:
    notation = str(product["notation"])
    price_unit = str(product["unit"])
    chf_factor = cost_adjustments.get("chf_to_eur_base_factor", 1.0)
    category = str(product["category"])
    cat_defaults = cost_adjustments.get("categories", {}).get(category, {})
    overrides = cost_adjustments.get("product_overrides", {}).get(notation, {})

    chf_material = product["material_cost"]
    material_cost = round(chf_material * chf_factor, 1) if chf_material is not None else None

    waste_pct = overrides.get("waste_pct", cat_defaults.get("waste_pct", 0.0))
    transport_pct = overrides.get("transport_pct", cat_defaults.get("transport_pct", 0.0))
    delivered = None
    if material_cost is not None:
        delivered = round(material_cost * (1 + waste_pct) * (1 + transport_pct), 1)

    price = {
        "material_cost": material_cost,
        "unit": price_unit,
        "waste_pct": waste_pct,
        "transport_pct": transport_pct,
        "delivered_material_cost": delivered,
    }

    labor_override = labor_overrides.get(notation, {})
    onsite = labor_override.get("onsite_hours_per_unit", product["labor_hours"])
    labor = {
        "onsite_hours_per_unit": onsite,
        "offsite_hours_per_unit": 0.0,
        "unit": price_unit,
    }
    if notation in cost_adjustments.get("prefab_notations", []):
        labor["offsite_hours_per_unit"] = labor_override.get("offsite_hours_per_unit", 0.0)
    elif "offsite_hours_per_unit" in labor_override:
        labor["offsite_hours_per_unit"] = labor_override["offsite_hours_per_unit"]

    demo_defaults = overrides.get("demolition", cat_defaults.get("demolition", {}))
    demolition = {
        "hours_per_unit": demo_defaults.get("hours_per_unit"),
        "disposal_cost": demo_defaults.get("disposal_cost"),
        "unit": price_unit,
    }
    return price, labor, demolition


def get_factor(factors: dict, kbob_id: str) -> dict | None:
    return factors.get(kbob_id_to_iri(kbob_id))


def layer_qty_per_parent_m2(layer: dict, defaults: dict) -> float:
    rule = layer.get("quantity_rule", "")
    if "timber_depth_m * 450" in rule:
        return layer.get("default_timber_depth_m", 0.16) * 450
    if "topping_thickness_m * 2500" in rule:
        return layer.get("default_topping_thickness_m", 0.08) * 2500
    if "steel_deck_mass_kg_m2" in rule:
        return layer.get("default_steel_deck_mass_kg_m2", 8)
    if "clay_infill_thickness_m * 1800" in rule:
        return layer.get("default_clay_infill_thickness_m", 0.1) * 1800
    if "insulation_thickness_m * 20" in rule:
        return layer.get("default_insulation_thickness_m", 0.14) * 20
    if "render_mass_kg_m2" in rule:
        return layer.get("default_render_mass_kg_m2", 25)
    if rule.endswith("opening_area_m2"):
        return defaults.get("window_frame_area_ratio", 1.0)
    if rule.endswith("glazing_area_m2"):
        return defaults.get("default_glazing_area_ratio", 0.75)
    return 1.0


def convert_gwp_to_price_unit(
    gwp_per_ref: float,
    ref_unit: str,
    price_unit: str,
    factor: dict,
    notation: str,
    unit_defaults: dict,
) -> dict[str, object]:
    ref_unit = normalize_unit(ref_unit)
    price_unit = normalize_unit(price_unit)
    if ref_unit == price_unit:
        return {"gwp_kg_co2eq": gwp_per_ref, "method": "direct"}

    density = factor.get("density")
    if price_unit == "m3" and ref_unit == "kg" and density:
        return {
            "gwp_kg_co2eq": gwp_per_ref * density,
            "method": "density",
            "density_kg_m3": density,
        }

    if price_unit == "m2" and ref_unit == "kg":
        area_defaults = unit_defaults.get("area_yield_kg_per_m2", {})
        thickness_defaults = unit_defaults.get("area_yield_thickness_m", {})
        kg_per_m2 = area_defaults.get(notation)
        thickness_m = thickness_defaults.get(notation)
        if kg_per_m2 is None and thickness_m is not None and density:
            kg_per_m2 = thickness_m * density
        if kg_per_m2 is not None:
            result: dict[str, object] = {
                "gwp_kg_co2eq": gwp_per_ref * kg_per_m2,
                "method": "area_yield",
                "mass_per_area_kg_m2": kg_per_m2,
            }
            if thickness_m is not None:
                result["thickness_m"] = thickness_m
            if density:
                result["density_kg_m3"] = density
            return result
        return {
            "gwp_kg_co2eq": gwp_per_ref,
            "method": "unconverted",
            "assumptions": f"missing mass_per_area_kg_m2 for {notation} (ref_unit={ref_unit}, price_unit={price_unit})",
        }

    if price_unit == "pcs" and ref_unit == "m2":
        m2_per_piece = unit_defaults.get("default_m2_per_piece", {}).get(notation, 2.0)
        return {
            "gwp_kg_co2eq": gwp_per_ref * m2_per_piece,
            "method": "piece_area",
            "area_per_piece_m2": m2_per_piece,
        }

    if price_unit == "m2" and ref_unit == "m2":
        return {"gwp_kg_co2eq": gwp_per_ref, "method": "direct"}

    return {
        "gwp_kg_co2eq": gwp_per_ref,
        "method": "unconverted",
        "assumptions": f"ref_unit={ref_unit}, price_unit={price_unit}",
    }


def compute_carbon(
    notation: str,
    price_unit: str,
    mappings: dict[str, dict[str, list[str]]],
    factors: dict,
    recipes: dict,
    unit_defaults: dict,
    kbob_labels: dict[str, str],
) -> dict:
    lookup_notation = NOTATION_ALIASES.get(notation, notation)
    mapping = mappings.get(lookup_notation, {"closeMatch": [], "relatedMatch": []})

    if lookup_notation in recipes.get("recipes", {}):
        recipe = recipes["recipes"][lookup_notation]
        components = []
        total_gwp = 0.0
        for layer in recipe["layers"]:
            kbob_id = layer["kbob_id"]
            factor = get_factor(factors, kbob_id)
            if not factor:
                continue
            qty = layer_qty_per_parent_m2(layer, unit_defaults)
            gwp = factor["gwp_total_kg_co2eq"] * qty
            total_gwp += gwp
            components.append(
                {
                    "notation": layer.get("notation"),
                    "kbob_id": kbob_id,
                    "quantity_per_price_unit": qty,
                    "quantity_unit": layer.get("quantity_unit"),
                    "gwp_kg_co2eq": round(gwp, 4),
                }
            )
        rows = []
        for layer in recipe["layers"]:
            kbob_id = layer["kbob_id"]
            factor = get_factor(factors, kbob_id)
            if factor:
                rows.append(build_row(kbob_id, factor, kbob_labels))
        return {
            "primary_match_type": "layer_recipe",
            "rows": rows,
            "carbon_per_price_unit": {
                "gwp_kg_co2eq": round(total_gwp, 4),
                "method": "layer_recipe_sum",
                "assumptions": recipe.get("quantity_basis"),
                "components": components,
            },
        }

    kbob_ids = mapping["closeMatch"] or mapping["relatedMatch"]
    if not kbob_ids:
        return {
            "primary_match_type": None,
            "rows": [],
            "carbon_per_price_unit": {"gwp_kg_co2eq": None, "method": "unmapped", "assumptions": None},
        }

    primary_type = "closeMatch" if mapping["closeMatch"] else "relatedMatch"
    primary_id = kbob_ids[0]
    factor = get_factor(factors, primary_id)
    if not factor:
        return {
            "primary_match_type": primary_type,
            "rows": [],
            "carbon_per_price_unit": {"gwp_kg_co2eq": None, "method": "unmapped", "assumptions": None},
        }

    gwp_result = convert_gwp_to_price_unit(
        factor["gwp_total_kg_co2eq"],
        factor.get("ref_unit", ""),
        price_unit,
        factor,
        lookup_notation,
        unit_defaults,
    )
    rows = [build_row(kid, get_factor(factors, kid), kbob_labels) for kid in kbob_ids if get_factor(factors, kid)]
    carbon = {"gwp_kg_co2eq": round(float(gwp_result["gwp_kg_co2eq"]), 4), "method": gwp_result["method"]}
    for key in ("density_kg_m3", "mass_per_area_kg_m2", "area_per_piece_m2", "thickness_m", "assumptions"):
        if gwp_result.get(key) is not None:
            carbon[key] = gwp_result[key]
    return {
        "primary_match_type": primary_type,
        "rows": rows,
        "carbon_per_price_unit": carbon,
    }


def build_row(kbob_id: str, factor: dict | None, kbob_labels: dict[str, str]) -> dict:
    if not factor:
        return {"kbob_id": kbob_id}
    return {
        "kbob_id": kbob_id,
        "iri": kbob_id_to_iri(kbob_id),
        "label_de": kbob_labels.get(kbob_id),
        "ref_unit": factor.get("ref_unit"),
        "gwp_total_kg_co2eq": factor.get("gwp_total_kg_co2eq"),
        "pe_total_kwh_oil_eq": factor.get("pe_total_kwh_oil_eq"),
        "ubp_total": factor.get("ubp_total"),
    }


def format_kbob_reference(kbob: dict) -> dict | None:
    rows = kbob.get("rows", [])
    carbon = kbob.get("carbon_per_price_unit", {})
    benchmark_uris = [row["iri"] for row in rows if row.get("iri")]
    primary = kbob.get("primary_match_type")
    method = carbon.get("method")
    if not benchmark_uris and not method:
        return None
    ref: dict[str, object] = {}
    if primary:
        ref["primary_match_type"] = primary
    if benchmark_uris:
        ref["benchmark_uris"] = benchmark_uris
    if method:
        ref["carbon_method"] = method
    return ref


def format_carbon_estimate(kbob: dict) -> dict | None:
    carbon = kbob.get("carbon_per_price_unit", {})
    method = carbon.get("method")
    if not method:
        return None
    estimate: dict[str, object] = {"method": method}
    if carbon.get("gwp_kg_co2eq") is not None:
        estimate["gwp_kg_co2eq"] = carbon["gwp_kg_co2eq"]
    for key in ("density_kg_m3", "mass_per_area_kg_m2", "area_per_piece_m2", "thickness_m", "assumptions"):
        if carbon.get(key) is not None:
            estimate[key] = carbon[key]
    if carbon.get("components"):
        estimate["components"] = carbon["components"]
    return estimate


def format_unit_price_entry(
    notation: str,
    price: dict,
    labor: dict,
    demolition: dict,
    kbob: dict,
    unc_cfg: dict,
    provenance_status: str,
) -> dict:
    entry: dict[str, object] = {
        "product": notation,
        "product_uri": notation_to_iri(notation),
        "material_cost": price["material_cost"],
        "price_unit": price["unit"],
        "waste_pct": price["waste_pct"],
        "transport_pct": price["transport_pct"],
        "onsite_labor_hours_per_unit": labor["onsite_hours_per_unit"],
    }
    if price.get("delivered_material_cost") is not None:
        entry["delivered_material_cost"] = price["delivered_material_cost"]
    offsite = labor.get("offsite_hours_per_unit") or 0.0
    if offsite:
        entry["offsite_labor_hours_per_unit"] = offsite
    demo = format_demolition(demolition)
    if demo:
        entry["demolition"] = demo
    kbob_ref = format_kbob_reference(kbob)
    if kbob_ref:
        entry["kbob"] = kbob_ref
    carbon = format_carbon_estimate(kbob)
    if carbon:
        entry["carbon_per_price_unit"] = carbon
    entry["uncertainty"] = build_uncertainty(price, labor, unc_cfg)
    entry["provenance_status"] = provenance_status
    return entry


def build_regional_benchmarks(
    regions_raw: list[dict[str, object]],
    material_benchmarks: dict,
    offsite_rates: dict[str, float | None],
) -> list[dict[str, object]]:
    material_by_code = {str(row["code"]): row for row in material_benchmarks.get("regions", [])}
    merged: list[dict[str, object]] = []
    for region in regions_raw:
        code = str(region["code"])
        material = material_by_code.get(code, {})
        offsite = offsite_rates.get(code)
        entry: dict[str, object] = {
            "code": code,
            "currency": region["currency"],
            "material_factor": material.get("material_factor", 1.0),
            "fx_to_eur": material.get("fx_to_eur", 1.0),
            "labor_factor": 1.0,
            "labor_unit_price": region["hourly_rate"],
        }
        if offsite is not None:
            entry["offsite_labor_unit_price"] = offsite
        merged.append(entry)
    return merged


def refresh_entry_carbon(
    entry: dict,
    *,
    mappings: dict,
    factors: dict,
    recipes: dict,
    unit_defaults: dict,
    kbob_labels: dict[str, str],
) -> list[str]:
    notation = str(entry["product"])
    price_unit = str(entry["price_unit"])
    kbob = compute_carbon(notation, price_unit, mappings, factors, recipes, unit_defaults, kbob_labels)
    carbon = format_carbon_estimate(kbob)
    if carbon:
        entry["carbon_per_price_unit"] = carbon
    kbob_ref = format_kbob_reference(kbob)
    if kbob_ref:
        entry["kbob"] = kbob_ref

    warnings: list[str] = []
    method = (carbon or {}).get("method")
    if method == "unmapped":
        warnings.append(f"unmapped: {notation}")
    elif method == "unconverted":
        warnings.append(f"unit mismatch: {notation} ({(carbon or {}).get('assumptions')})")
    return warnings


def refresh_entry_uncertainty(
    entry: dict,
    cost_adjustments: dict,
) -> None:
    notation = str(entry["product"])
    category = infer_product_category(notation)
    unc_cfg = resolve_uncertainty_config(notation, category, cost_adjustments)
    price = {
        "material_cost": entry.get("material_cost"),
        "delivered_material_cost": entry.get("delivered_material_cost"),
    }
    labor = {
        "onsite_hours_per_unit": entry.get("onsite_labor_hours_per_unit"),
        "offsite_hours_per_unit": entry.get("offsite_labor_hours_per_unit", 0.0),
    }
    entry["uncertainty"] = build_uncertainty(price, labor, unc_cfg)
    if isinstance(entry.get("uncertainty"), dict):
        entry["uncertainty"].pop("confidence", None)
    entry["provenance_status"] = resolve_provenance_status(notation, cost_adjustments)


def refresh_products_uncertainty(products_path: Path) -> int:
    cost_adjustments = json.loads(COST_ADJUSTMENTS_PATH.read_text(encoding="utf-8"))
    products_path = products_path.resolve()
    products_doc = json.loads(products_path.read_text(encoding="utf-8"))
    entries = products_doc.get("entries", {})
    if not isinstance(entries, dict):
        raise SystemExit("Expected entries dict in baseline products JSON")

    for entry in entries.values():
        refresh_entry_uncertainty(entry, cost_adjustments)

    products_path.write_text(json.dumps(products_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Refreshed uncertainty for {len(entries)} entries in {products_path.relative_to(REPO_ROOT)}")
    return 0


def recompute_products_carbon(products_path: Path) -> int:
    lca_data = json.loads(LCA_FACTORS_PATH.read_text(encoding="utf-8"))
    factors = lca_data["factors"]
    recipes = json.loads(LAYER_RECIPES_PATH.read_text(encoding="utf-8"))
    unit_defaults = json.loads(UNIT_DEFAULTS_PATH.read_text(encoding="utf-8"))
    mappings = load_mappings()
    kbob_labels = load_kbob_labels()

    products_path = products_path.resolve()
    products_doc = json.loads(products_path.read_text(encoding="utf-8"))
    entries = products_doc.get("entries", {})
    if not isinstance(entries, dict):
        raise SystemExit("Expected entries dict in baseline products JSON")

    warnings: list[str] = []
    for notation, entry in entries.items():
        warnings.extend(
            refresh_entry_carbon(
                entry,
                mappings=mappings,
                factors=factors,
                recipes=recipes,
                unit_defaults=unit_defaults,
                kbob_labels=kbob_labels,
            )
        )

    products_path.write_text(json.dumps(products_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Recomputed carbon for {len(entries)} entries in {products_path.relative_to(REPO_ROOT)}")
    if warnings:
        print("Warnings:", file=sys.stderr)
        for warning in warnings:
            print(f"  - {warning}", file=sys.stderr)
        return 1
    return 0


def main() -> int:
    lca_data = json.loads(LCA_FACTORS_PATH.read_text(encoding="utf-8"))
    factors = lca_data["factors"]
    recipes = json.loads(LAYER_RECIPES_PATH.read_text(encoding="utf-8"))
    unit_defaults = json.loads(UNIT_DEFAULTS_PATH.read_text(encoding="utf-8"))
    mappings = load_mappings()
    kbob_labels = load_kbob_labels()
    products_raw = load_excel_products()
    regions_raw = load_excel_regions()

    cost_adjustments = json.loads(COST_ADJUSTMENTS_PATH.read_text(encoding="utf-8"))
    labor_overrides = {
        str(item["notation"]): item
        for item in json.loads(LABOR_OVERRIDES_PATH.read_text(encoding="utf-8")).get("products", [])
    }
    offsite_rates = {str(r["code"]): r.get("offsite_hourly_rate") for r in cost_adjustments.get("regions", [])}

    material_benchmarks = json.loads(REGIONAL_MATERIAL_PATH.read_text(encoding="utf-8"))
    material_codes = {str(r["code"]) for r in material_benchmarks.get("regions", [])}
    labor_codes = {str(r["code"]) for r in regions_raw}

    warnings: list[str] = []
    missing_material = labor_codes - material_codes
    if missing_material:
        warnings.append(
            f"regional-material-benchmarks.json missing material_factor for: {sorted(missing_material)}"
        )

    entries_out: dict[str, dict] = {}

    for product in products_raw:
        notation = str(product["notation"])
        price_unit = str(product["unit"])
        kbob = compute_carbon(notation, price_unit, mappings, factors, recipes, unit_defaults, kbob_labels)
        if kbob["carbon_per_price_unit"]["method"] == "unmapped":
            warnings.append(f"unmapped: {notation}")
        elif kbob["carbon_per_price_unit"]["method"] == "unconverted":
            warnings.append(f"unit mismatch: {notation} ({kbob['carbon_per_price_unit']['assumptions']})")

        price, labor, demolition = build_price_labor_demolition(product, cost_adjustments, labor_overrides)
        unc_cfg = resolve_uncertainty_config(notation, str(product["category"]), cost_adjustments)
        provenance_status = resolve_provenance_status(notation, cost_adjustments)
        entries_out[notation] = format_unit_price_entry(
            notation, price, labor, demolition, kbob, unc_cfg, provenance_status
        )

    issued = cost_adjustments.get("issued", "2026-07-01")
    source = str(XLSX_PATH.relative_to(REPO_ROOT))

    products_doc = {
        "version": "v12",
        "issued": issued,
        "anchor_region": ANCHOR_REGION,
        "source": source,
        "license": LICENSE,
        "disclaimer": DISCLAIMER,
        "carbon_source": lca_data.get("source"),
        "carbon_source_version": lca_data.get("version"),
        "entries": entries_out,
    }

    regional_merged = build_regional_benchmarks(regions_raw, material_benchmarks, offsite_rates)
    regions_out = {str(row["code"]): row for row in regional_merged}
    regional_doc = {
        "version": "v12",
        "issued": issued,
        "anchor_region": material_benchmarks.get("anchor", ANCHOR_REGION),
        "source": source,
        "license": LICENSE,
        "disclaimer": DISCLAIMER,
        "regions": regions_out,
    }

    labor_legacy = []
    for region in regional_merged:
        labor_legacy.append(
            {
                "code": region["code"],
                "hourly_rate": region["labor_unit_price"],
                "currency": region["currency"],
                "offsite_hourly_rate": region.get("offsite_labor_unit_price"),
            }
        )

    material_legacy = {
        "version": "v12",
        "issued": issued,
        "anchor": material_benchmarks.get("anchor", ANCHOR_REGION),
        "license": LICENSE,
        "disclaimer": DISCLAIMER,
        "notes": material_benchmarks.get("notes"),
        "regions": material_benchmarks.get("regions", []),
    }

    OUT_PRODUCTS.parent.mkdir(parents=True, exist_ok=True)
    OUT_PRODUCTS.write_text(json.dumps(products_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    OUT_REGIONAL.write_text(json.dumps(regional_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    OUT_REGIONAL_LABOR.write_text(
        json.dumps(
            {
                "version": "v12",
                "source": source,
                "license": LICENSE,
                "disclaimer": DISCLAIMER,
                "issued": issued,
                "regions": labor_legacy,
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    OUT_REGIONAL_MATERIAL.write_text(json.dumps(material_legacy, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Wrote {len(entries_out)} entries to {OUT_PRODUCTS.relative_to(REPO_ROOT)}")
    print(f"Wrote {len(regional_merged)} regions to {OUT_REGIONAL.relative_to(REPO_ROOT)}")
    if warnings:
        print("Warnings:", file=sys.stderr)
        for w in warnings:
            print(f"  - {w}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--recompute-carbon":
        target = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT_PRODUCTS
        raise SystemExit(recompute_products_carbon(target))
    if len(sys.argv) > 1 and sys.argv[1] == "--refresh-uncertainty":
        target = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT_PRODUCTS
        raise SystemExit(refresh_products_uncertainty(target))
    raise SystemExit(main())
