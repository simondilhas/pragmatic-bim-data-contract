#!/usr/bin/env python3
"""Convert KBOB ecobilans construction Excel to structured JSON for easier parsing."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    REPO_ROOT
    / "temp/Oekobilanzdaten_ Baubereich_Donne_ecobilans_construction_2009-1-2022_v8.02.xlsx"
)
DEFAULT_JSON = REPO_ROOT / "temp/kbob-ecobilans-construction-v8.02.json"

ENTRY_ID_PATTERN = re.compile(r"^\d{2}\.\d{3}$")
GROUP_ID_PATTERN = re.compile(r"^\d{1,2}$")
VERSION_PATTERN = re.compile(r"Version\s+([\d.]+)", re.IGNORECASE)

# Fixed semantic keys for leading identity columns shared across data sheets.
IDENTITY_COLUMNS: dict[int, str] = {
    0: "id",
    1: "uuid",
    2: "label_de",
    3: "disposal_id",
    4: "disposal_label_de",
    5: "density",
    6: "ref_unit",
}

DATA_SHEETS = (
    "Baumaterialien Matériaux",
    "Gebäudetechnik Technique",
    "Energie Énergie",
    "Transporte Transports",
    "Entsorgung Déchets",
)

SHEET_KEYS = {
    "Baumaterialien Matériaux": "baumaterialien",
    "Gebäudetechnik Technique": "gebaeudetechnik",
    "Energie Énergie": "energie",
    "Transporte Transports": "transporte",
    "Entsorgung Déchets": "entsorgung",
}

# Per-sheet trailing label columns (French translations).
SHEET_TRAILING_COLUMNS: dict[str, dict[int, str]] = {
    "Baumaterialien Matériaux": {29: "label_fr", 30: "disposal_label_fr"},
    "Gebäudetechnik Technique": {23: "label_fr", 24: "disposal_label_fr"},
    "Energie Énergie": {12: "label_fr"},
    "Transporte Transports": {25: "label_fr"},
    "Entsorgung Déchets": {19: "label_fr"},
}

# Per-sheet identity columns (default covers Baumaterialien, Gebäudetechnik, Entsorgung).
SHEET_IDENTITY_COLUMNS: dict[str, dict[int, str]] = {
    "Energie Énergie": {
        0: "id",
        1: "uuid",
        2: "label_de",
        3: "reference_basis",
        4: "unit",
    },
    "Transporte Transports": {
        0: "id",
        1: "uuid",
        2: "label_de",
        3: "reference_basis",
        4: "unit",
    },
    "Entsorgung Déchets": {
        0: "id",
        1: "uuid",
        2: "label_de",
        3: "ref_unit",
    },
}

TEXT_IDENTITY_KEYS = {
    "id",
    "uuid",
    "label_de",
    "disposal_id",
    "disposal_label_de",
    "label_fr",
    "disposal_label_fr",
    "reference_basis",
    "unit",
    "ref_unit",
}

UNIT_TOKENS = {"UBP", "kWh oil-eq", "kg CO2-eq", "kg C"}


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == "-":
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None



def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\n", " ").strip()
    return text or None


def primary_header_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.split("\n")[0].strip()


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "value"


def is_french_header(value: str) -> bool:
    french_markers = (
        "fabrication",
        "elimination",
        "renouvelable",
        "non renouvelable",
        "emissions de gaz",
        "contenu dans le produit",
        "utilise sous forme",
        "expl",
        "vehicule",
        "infrastructure",
        "dimension",
        "combustibles",
        "materiaux de construct",
        "technique du batiment",
        "traitement des dechets",
        "transports",
    )
    lowered = slugify(value)
    return any(marker in lowered for marker in french_markers)


def is_note_text(value: str) -> bool:
    return value.startswith("Hinweis:") or value.startswith("Remarque:")


def classify_row_id(raw_id: object) -> tuple[str | None, str | None]:
    if raw_id is None:
        return None, None
    if isinstance(raw_id, (int, float)) and not isinstance(raw_id, bool):
        if float(raw_id).is_integer():
            text = str(int(raw_id))
        else:
            text = str(raw_id).strip()
    else:
        text = str(raw_id).strip()
    if not text:
        return None, None
    if ENTRY_ID_PATTERN.match(text):
        return "entry", text
    if GROUP_ID_PATTERN.match(text):
        return "group", text
    return "other", text


def find_header_rows(rows: list[tuple[object, ...]]) -> tuple[int, int, int]:
    """Return (id_row_idx, unit_row_idx, max_col) using zero-based row indices."""
    id_row_idx = next(
        (idx for idx, row in enumerate(rows) if normalize_text(row[0]) == "ID-Nummer"),
        None,
    )
    if id_row_idx is None:
        raise ValueError("Could not locate header row containing 'ID-Nummer'")

    unit_row_idx = next(
        (
            idx
            for idx in range(id_row_idx, min(id_row_idx + 8, len(rows)))
            if any(normalize_text(cell) in UNIT_TOKENS for cell in rows[idx])
        ),
        id_row_idx + 5,
    )

    max_col = 0
    for row in rows[id_row_idx : unit_row_idx + 1]:
        for idx, cell in enumerate(row):
            if normalize_text(cell):
                max_col = max(max_col, idx)
    return id_row_idx, unit_row_idx, max_col + 1


def forward_fill_row(row: tuple[object, ...], max_col: int) -> list[str | None]:
    filled: list[str | None] = []
    current: str | None = None
    for col_idx in range(max_col):
        cell = normalize_text(row[col_idx] if col_idx < len(row) else None)
        if cell and not is_note_text(cell) and not is_french_header(cell):
            current = cell
        filled.append(current)
    return filled


def nearest_top_header(top_row: tuple[object, ...], col_idx: int) -> str | None:
    for idx in range(col_idx, -1, -1):
        cell = primary_header_text(top_row[idx] if idx < len(top_row) else None)
        if cell and not is_note_text(cell):
            return cell
    return None


def identity_columns_for(sheet_name: str) -> dict[int, str]:
    if sheet_name in SHEET_IDENTITY_COLUMNS:
        return dict(SHEET_IDENTITY_COLUMNS[sheet_name])
    return dict(IDENTITY_COLUMNS)


def category_from_top_header(top: str | None, sub: str | None) -> str:
    top_slug = slugify(top or "")
    sub_slug = slugify(sub or "")

    if "ubp" in top_slug:
        return "ubp"
    if "treibhausgas" in top_slug or "thg" in top_slug:
        return "gwp"
    if "biogener" in top_slug:
        return "biogenic_carbon"
    if "primarenergie" in top_slug.replace("_", ""):
        if "am_standort" in sub_slug or "du_site" in sub_slug:
            return "pe_on_site_renewable"
        if "nicht_erneuerbar" in sub_slug or "graue_energie" in sub_slug:
            return "pe_non_renewable"
        if "erneuerbar" in sub_slug:
            return "pe_renewable"
        return "pe"
    if "pef" in top_slug:
        if "pef_ne" in top_slug or "pef_ne" in sub_slug:
            return "pe_non_renewable"
        if "pef_e" in top_slug or "pef_e" in sub_slug:
            return "pe_renewable"
        return "pe"
    return "value"


def phase_from_detail_headers(parts: list[str]) -> str:
    lowered = slugify(" ".join(parts))
    if "herstellung_energetisch" in lowered or "energetisch_genutzt" in lowered:
        return "manufacturing_energy"
    if "herstellung_stofflich" in lowered or "stofflich_genutzt" in lowered:
        return "manufacturing_material"
    if "herstellung" in lowered:
        return "manufacturing"
    if "entsorgung" in lowered:
        return "disposal"
    if "betrieb" in lowered:
        return "operation"
    if "fahrzeug" in lowered:
        return "vehicle"
    if "infrastruktur" in lowered:
        return "infrastructure"
    return "total"


def build_column_keys(
    header_rows: list[tuple[object, ...]],
    id_row_idx: int,
    unit_row_idx: int,
    max_col: int,
    sheet_name: str,
) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    used_keys: dict[str, int] = {}
    trailing = SHEET_TRAILING_COLUMNS.get(sheet_name, {})
    identity_columns = identity_columns_for(sheet_name)

    top_level = [nearest_top_header(header_rows[id_row_idx], col_idx) for col_idx in range(max_col)]
    sub_level = forward_fill_row(header_rows[id_row_idx + 1], max_col) if id_row_idx + 1 < len(header_rows) else [None] * max_col
    detail_rows = header_rows[id_row_idx + 2 : unit_row_idx]

    for col_idx in range(max_col):
        if col_idx in identity_columns:
            key = identity_columns[col_idx]
        elif col_idx in trailing:
            key = trailing[col_idx]
        else:
            detail_parts: list[str] = []
            for row in detail_rows:
                cell = normalize_text(row[col_idx] if col_idx < len(row) else None)
                if not cell or is_note_text(cell) or is_french_header(cell):
                    continue
                if cell not in detail_parts:
                    detail_parts.append(cell)
            category = category_from_top_header(top_level[col_idx], sub_level[col_idx])
            phase = phase_from_detail_headers(detail_parts)
            key = f"{category}_{phase}"
            parts = [part for part in (top_level[col_idx], sub_level[col_idx], *detail_parts) if part]

        count = used_keys.get(key, 0)
        used_keys[key] = count + 1
        if count:
            key = f"{key}_{count + 1}"

        unit = normalize_text(header_rows[unit_row_idx][col_idx] if col_idx < len(header_rows[unit_row_idx]) else None)
        headers = [
            normalize_text(header_rows[row_idx][col_idx] if col_idx < len(header_rows[row_idx]) else None)
            for row_idx in range(id_row_idx, unit_row_idx + 1)
        ]
        columns.append(
            {
                "index": col_idx,
                "key": key,
                "unit": unit if unit in UNIT_TOKENS else None,
                "headers": [item for item in headers if item],
            }
        )
    return columns


def row_to_record(
    row: tuple[object, ...],
    columns: list[dict[str, Any]],
    sheet_name: str,
) -> dict[str, Any] | None:
    raw_id = row[0] if row else None
    label_de = normalize_text(row[2] if len(row) > 2 else None)
    if raw_id is None and label_de is None:
        return None

    row_type, row_id = classify_row_id(raw_id)
    if row_type is None and label_de:
        row_type = "group"

    record: dict[str, Any] = {"row_type": row_type or "other"}
    if row_id is not None:
        record["id"] = row_id

    for column in columns:
        col_idx = column["index"]
        if col_idx >= len(row):
            continue
        value = row[col_idx]
        key = column["key"]

        if key in TEXT_IDENTITY_KEYS:
            text = normalize_text(value)
            if text is not None:
                record[key] = text
            continue

        if key == "density":
            parsed = parse_float(value)
            if parsed is not None:
                record[key] = parsed
            elif normalize_text(value) == "-":
                record[key] = None
            continue

        if row_type != "entry":
            continue

        parsed = parse_float(value)
        if parsed is None:
            continue
        record.setdefault("indicators", {})[key] = {
            "value": parsed,
            "unit": column["unit"],
        }

    if row_type == "group" and "label_de" in record:
        record.setdefault("label", record["label_de"])

    return record


def load_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    id_row_idx, unit_row_idx, max_col = find_header_rows(rows)
    data_start_row = unit_row_idx + 2  # one-based Excel row after unit row

    columns = build_column_keys(rows, id_row_idx, unit_row_idx, max_col, sheet_name)
    parsed_rows: list[dict[str, Any]] = []
    for row in rows[data_start_row - 1 :]:
        record = row_to_record(row, columns, sheet_name)
        if record:
            parsed_rows.append(record)

    return {
        "sheet_name": sheet_name,
        "data_start_row": data_start_row,
        "columns": columns,
        "rows": parsed_rows,
        "counts": {
            "total": len(parsed_rows),
            "groups": sum(1 for item in parsed_rows if item.get("row_type") == "group"),
            "entries": sum(1 for item in parsed_rows if item.get("row_type") == "entry"),
        },
    }


def extract_version(rows: list[tuple[object, ...]]) -> str | None:
    for row in rows[:3]:
        for cell in row:
            text = normalize_text(cell)
            if not text:
                continue
            match = VERSION_PATTERN.search(text)
            if match:
                return match.group(1)
    return None


def load_explanation_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    paragraphs: list[str] = []
    for row in rows:
        for cell in row:
            text = normalize_text(cell)
            if text and len(text) > 40 and text not in paragraphs:
                paragraphs.append(text)
    return {
        "sheet_name": ws.title,
        "notes": paragraphs,
        "version": extract_version(rows),
    }


def convert_workbook(xlsx_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    payload: dict[str, Any] = {
        "source": xlsx_path.name,
        "standard": "KBOB / ecobau / IPB 2009/1:2022",
        "version": None,
        "sheets": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.startswith("Erläuterung"):
            explanation = load_explanation_sheet(ws)
            payload["explanation"] = explanation
            payload["version"] = payload["version"] or explanation.get("version")
            continue
        if sheet_name not in DATA_SHEETS:
            continue
        sheet_payload = load_sheet(ws, sheet_name)
        payload["sheets"][SHEET_KEYS[sheet_name]] = sheet_payload

    wb.close()
    return payload


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert KBOB ecobilans construction Excel workbook to JSON."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Input .xlsx file (default: {DEFAULT_XLSX.relative_to(REPO_ROOT)})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_JSON,
        help=f"Output .json file (default: {DEFAULT_JSON.relative_to(REPO_ROOT)})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.is_file():
        raise SystemExit(f"Input file not found: {args.input}")

    payload = convert_workbook(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    total_entries = sum(sheet["counts"]["entries"] for sheet in payload["sheets"].values())
    print(f"Wrote {args.output}")
    print(f"  sheets: {len(payload['sheets'])}")
    print(f"  entries: {total_entries}")


if __name__ == "__main__":
    main()
